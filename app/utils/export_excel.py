"""Excel export utility using openpyxl."""

import io
from typing import Any


def export_to_excel(
    data: list[dict[str, Any]], columns: list[str], sheet_name: str = "Sheet1"
) -> bytes:
    """Convert list of dicts to Excel workbook bytes."""
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name

    # Header
    for col_idx, col_name in enumerate(columns, 1):
        sheet.cell(row=1, column=col_idx, value=col_name)

    # Data rows
    for row_idx, row in enumerate(data, 2):
        for col_idx, col_name in enumerate(columns, 1):
            sheet.cell(row=row_idx, column=col_idx, value=row.get(col_name, ""))

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()
